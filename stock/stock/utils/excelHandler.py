#!/usr/bin/env python
# coding=utf-8

import os
from copy import copy
import xlrd
import xlwt
import openpyxl
from logHandler import Jcl_Logger

"""
处理excel方法如下：
openpyxl,xlutils(xlrd,xlwt),pandas,xlwings,win32com,xlsxwriter,datanitro
处理excel有两种方法：需要用到xlrd模块或者openpyxl模块，区别：
1.xlrd能处理的最大行数为65535，超过这个行数的文件就需要用到openpyxl，openpyxl最大支持1048576行；
2.openpyxl不支持xls格式文件

excel的格式：
xlsm，xlsx,xltm,xltx

openpyxl
https://www.cnblogs.com/simple-free/p/9160685.html
https://www.cnblogs.com/fengf233/p/10880782.html
xlutils(xlrd,xlwt)
https://blog.csdn.net/catch_dreamer/article/details/103731100

pandas
https://www.cnblogs.com/liulinghua90/p/9935642.html
https://www.jianshu.com/p/840ba135df30
https://blog.csdn.net/aasdad1/article/details/91812714

"""


'''
dict['key']只能获取存在的值，如果不存在则触发KeyError
dict.get(key,default=None)，返回指定键的值，如果值不在字典中返回默认值None
excel文件中请求数据有可能为空，所以用get方法获取
'''

# excel文件放在固定的目录下
testdata_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'testdatafiles')

# excel类
class ExcelHandler():

    def __init__(self):
        self.excelPath = os.path.join(testdata_path,'excel')
        # print(self.excelPath)

    # 获取某文件目录下的所有excel文件
    def get_tables_list(self):
        excel_file_list = []
        filenames = os.listdir(self.excelPath)
        for filename in filenames:
            if filename.split('.')[-1] in ('xls','xlsx','xlsm'):
                excel_file_list.append(filename)
            else:
                pass
        return excel_file_list

    # 获取文件夹下所有excel下的所有的sheet中，输出最多行数的sheet的行数
    def get_all_tables_longest_rows_number(self):
        maxRowCount_list = []
        table_list = self.get_tables_list()
        print(table_list)
        try:
            for tablename in table_list:
                # 获取子sheet列表
                print(tablename)
                sheets = self.get_table_and_open(tablename).sheets()
                # 初始值假设为第一个表的行数
                maxRowCount = sheets[0].nrows
                for sheet in sheets:
                    if sheet.nrows !=0 and sheet.nrows>maxRowCount:
                        maxRowCount = sheet.nrows
                maxRowCount_list.append(maxRowCount)
            return max(maxRowCount_list)
        except Exception as e:
            print(e)

    # 在excel文件夹下寻找某个名称的excel表格
    def get_table_and_open(self,tablename):
        # 遍历读取excel文件下的各个excel，便于拓展，excel文件夹下可能会有多个excel文件
        filenames = os.listdir(self.excelPath)
        for filename in filenames:
            if filename.split('.')[-1] in ('xls','xlsx','xlsm'):
                table = xlrd.open_workbook(os.path.join(self.excelPath,tablename))
                break
            else:
                Jcl_Logger('jcllogger').getlog().debug("这个excel不存在，请检查参数tablename")
        return table


    # 获取某个表格子sheet个数
    def get_sheets_number(self,tablename):
        table = self.get_table_and_open(tablename)
        # 获取表的子sheet列表
        sheets = table.sheets()
        # 列表长度就是子sheet个数
        self.sheetCount = len(sheets)

    def get_all_sheets_longest_rows_number(self,tablename):
        # 获取子sheet列表
        sheets = self.get_table_and_open(tablename).sheets()
        # 初始值假设为第一个表的行数
        maxRowCount = sheets[0].nrows
        for sheet in sheets:
            if sheet.nrows !=0 and sheet.nrows>maxRowCount:
                maxRowCount = sheet.nrows
        return maxRowCount

    # 读取某个子表中的所有行的取值，存为列表，一行内容为一个元素；
    def readRows(self,tablename,sheetIndex=0):
        table = self.get_table_and_open(tablename)
        # 通过传入的sheetIndex确定是要打开哪个子sheet
        sheet = table.sheet_by_index(sheetIndex)
        #row_values(r)表示第r行所有的行单元格里面的取值
        return [sheet.row_values(r) for r in range(0,sheet.nrows)]

    # 读取某个子表中的所有列的取值，一列为一个元素
    def readCols(self,tablename,sheetIndex=0):
        table = self.get_table_and_open(tablename)
        sheet = table.sheet_by_index(sheetIndex)
        # col_values(c)表示第r行所有的行单元格里面的取值
        return [sheet.col_values(c) for c in range(0,sheet.ncols)]

    # 返回一个空行，为下面读取某行功能使用
    def readABlankRow(self,tablename,sheetIndex=0):
        BlankRow=[]
        sheet = self.get_table_and_open(tablename).sheet_by_index(sheetIndex)
        try:
            for c in range(0,sheet.ncols):
                # 输出一个空行，需要按照列的长度一个单元格一个单元格的输出
                BlankRow.append("")
            return BlankRow
        except Exception as e:
            print(e)
            return None

    # 读取某行
    def readOneRow(self,tablename,rowIndex,sheetIndex=0):
        try:
            table = self.get_table_and_open(tablename)
            sheet = table.sheet_by_index(sheetIndex)
            #只要这一行有一个地方有值，都能返回数据，除非这行都是空的，才会异常，走下面的逻辑
            return sheet.readRows(sheetIndex)[rowIndex]
        except BaseException as e:
            print(e)
            #那就是设置一个空行
            return self.readABlankRow(sheetIndex)

    # 返回一个空列
    def readABlankCol(self,tablename,sheetIndex=0):
        BlankCol = []
        table = self.get_table_and_open(tablename)
        sheet = table.sheet_by_index(sheetIndex)
        for r in range(0,sheet.nrows):
            #输出某列的值，需要按照行的长度一个一个的输出来一列
            BlankCol.append("")
        return BlankCol

    # 读取某列
    def readOneCol(self,tablename,colIndex,sheetIndex=0):
        try:
            table = self.get_table_and_open(tablename)
            sheet = table.sheet_by_index(sheetIndex)
            #这一列只要不都是空，就能返回数据，除非此列都是空的
            return sheet.readCols(sheetIndex)[colIndex]
        except Exception as e:
            print(e)
            #都是空的，走这的逻辑
            return self.readABlankCol(sheetIndex)

    # 读取单元格
    def readCell(self,tablename,rowIndex,colIndex,sheetIndex=0):
        try:
            return self.readRows(tablename,sheetIndex)[rowIndex][colIndex]
        except Exception as e:
            print(e)
            return ""

    # 读取首行标题行：对于二维表格，一般首行为标题
    def readTitleRow(self,tablename,sheetIndex=0):
        try:
            return self.readRows(tablename,sheetIndex)[0]
        except Exception as e:
            print(e)
            return self.readABlankRow(tablename,sheetIndex)

    # 读取除去第一行剩下的所有行
    def readDataExceptTitleRow(self,tablename,sheetIndex=0):
        try:
            maxRowIndex = len(self.readRows(tablename,sheetIndex))
            return [self.readRows(tablename,sheetIndex)[i] for i in range(1, maxRowIndex)]
        except Exception as e:
            print(e)
            return self.readABlankRow(tablename,sheetIndex)

    # 读取首列标题列：对于二维表格，一般首列为id号
    def readTitleCol(self,tablename,sheetIndex=0):
        try:
            return self.readCols(tablename,sheetIndex)[0]
        except Exception as e:
            print(e)
            return self.readABlankCol(tablename,sheetIndex)

    # 读取除去第一列剩下的所有列
    def readDataExceptTitleCol(self, tablename,sheetIndex=0):
        try:
            maxColIndex = len(self.readCols(tablename,sheetIndex))
            return [self.readCols(tablename,sheetIndex)[i] for i in range(1, maxColIndex)]
        except Exception as e:
            print(e)
            return self.readABlankCol(tablename,sheetIndex)

    # 读取除了第一行和第一列的数据(行优先存储,先把第一列排除，剩余的内容中再把第一列排除)
    def readDataExceptTitleRowAndFirstCol(self,tablename,sheetIndex=0):
        try:
            # 计算出行数
            maxRowIndex = len(self.readRows(tablename,sheetIndex))
            # 读取剩余行的内容
            exceptRows = [self.readRows(tablename,sheetIndex)[i] for i in range(1,maxRowIndex)]
            datas = []
            # row为每一行的内容，exceptRows为所有行的内容的列表
            for row in exceptRows:
                rowDatas = []
                # 某一行的长度，就是整个列表的列数
                maxColIndex = len(row)
                for colIndex in range(1,maxColIndex):
                    # 第row行的第2,3,4,5......的数据
                    rowDatas.append(row[colIndex])
                datas.append(rowDatas)
            return datas
        except Exception as e:
            print(e)
            return None

    # 向excel中写入数据
    def writeToExcel(self,xlsfile,rowIndex,colIndex,sheetIndex=0):
        table = xlrd.open_workbook(xlsfile)
        sheet = table.sheet_by_index(sheetIndex)
        cell_value = sheet.cell(rowIndex,colIndex)
        print(cell_value)
        # 保存xlsfile
        wb = copy(table)
        wb.save(xlsfile)





# 待整理
excel_path = os.path.join(os.getcwd(),'python_excel.xlsx')

data = []

def read_xlsx_by_xlrd(excel_path):
    # 打开excel文件
    xlsx_data = xlrd.open_workbook(excel_path)
    # 进入第一张表,下面3条效果相同
    sheet_name1 = xlsx_data.sheets()[0]
    sheet_name2 = xlsx_data.sheet_by_index(0)
    sheet_name3 = xlsx_data.sheet_by_name("jcl0709001")

    # 获取所有表的索引,执行会提示'Book' object has no attribute 'sheet'
    # all_sheet_index = xlsx_data.sheet()
    # print(all_sheet_index)

    # 获取所有表的索引？
    all_sheet_index = xlsx_data.sheets()
    # print(all_sheet_index)

    # 获取总行数
    total_nrows = sheet_name1.nrows

    # 获取总列数
    total_ncols = sheet_name1.ncols

    # 获取首行,行数据，作为下面字典的key
    line_value = sheet_name1.row_values(0)
    # print(line_value)

    # 获取首列，列数据
    col_value = sheet_name1.col_values(0)
    # print(col_value)

    # 读取除了 第一行和第一列的数据
    for i in range(1, total_nrows):
        data_row = {}
        for j in range(0, total_ncols):
            # 获取指定单元格数据：sheet.cell(a,b).value  a表示行的下标值，b表示列的下标值；两者可以调换位置；
            # 注意：不加.value取出的值前面会带个text的东西
            # sheet_name1.row_values(0)作为字典的key
            data_row[sheet_name1.row_values(0)[j]] = sheet_name1.cell(i, j).value  # 根据行数来取对应列的值，并添加到字典中
            # print(data_row)
        data.append(data_row)
    return data


def read_xlsx_by_openpyxl(excel_path):
    # 打开excel文件,openpyxl不支持xls文件
    xlsx_data = openpyxl.load_workbook(excel_path)

    # 获取所有表格的名字
    sheet_names = xlsx_data.sheetnames
    print('所有表的名字：',sheet_names)

    # 获取处于active的表格
    active_sheet1 = xlsx_data.active
    print(active_sheet1)
    active_sheet2 = xlsx_data.get_active_sheet()
    print(active_sheet2)

    # 删除一个WorkSheet, 注意是WorkSheet对象, 不是名字
    # remove(worksheet)

    # 保存到文件, 记住有写入操作记得保存!filename为目标文件夹
    # save(filename)

    # 创建一个WorkSheet，可传title和index两个参数,不传生成的WorkSheet名在'Sheet'后面递增加数字
    new_sheet = xlsx_data.create_sheet(title='jclnewsheet0709',index=6)
    print(new_sheet)


    # 根据表格名获取表格的数据
    sheet_data1 = xlsx_data.get_sheet_by_name('jcl0709001')
    print('某个表的数据：',sheet_data1)
    sheet_data2 = xlsx_data['jcl0709001']
    print('某个表的数据：',sheet_data2)

    # 返回所有单元格的值，是个列表形式
    all_unit_values = sheet_data1.values
    print(all_unit_values)

    # 读所有单元格数据
    all_unit_values2 = sheet_data1.get_cell_collection()
    print(all_unit_values2)

    # 获得多个单元格，返回一个生成器
    iter_rows = sheet_data1.iter_rows(range_string=None, row_offset=0, column_offset=0)
    print(iter_rows)

    # 获取表的标题
    sheet_title = sheet_data1.title
    print(sheet_title)

    # 返回所有行，所有列
    all_rows = sheet_data1.rows
    all_cols = sheet_data1.columns

    # 获取最大/小行数
    max_row = sheet_data1.max_row
    print('最大行数：',max_row)
    min_row = sheet_data1.min_row
    print('最小行数：',min_row)

    # 获取最大/小列数
    max_col = sheet_data1.max_column
    print('最大列数：',max_col)
    min_col = sheet_data1.min_column
    print('最小列数：',min_col)

    # 获取某个单元格的值，列在前行在后，引号印起来，其中列ABCDEFGHIJKLMN。。。。 行123456。。。。
    some_value_1 = sheet_data1['B1'].value
    print('1某行某列的值为:',some_value_1)

    # 获取a行，b列的值，a，b其中一个必须为1或以上，Cell的row和column都是从1开始的！！！
    some_value_2 = sheet_data1.cell(row=1,column=1)
    print('2某行某列的值为:',some_value_2.value)

    # 设置单元格的值
    some_value_2.value = 100


    # 获取多个,返回多行数据,类型为tuple
    multi_data=sheet_data1['A1:A2']
    print(multi_data)





if __name__=='__main__':
    Jcl_Logger('jcllogger').getlog().debug("开始执行测试代码")

    excel=ExcelHandler()

    # 读取文件夹中某个excel表格
    print(excel.get_table_and_open('interface.xls'))

    # 读取某个子sheet的全部行内容
    # print(excel.readRows('interface.xls',0))

    # 读取某个子sheet的全部列内容
    # print(excel.readCols('interface.xls',0))

    # 读取某表格中的一个单元格的值
    # print(excel.readCell('interface.xls',0, 1,sheetIndex=0))
    # print(excel.readCell('interface.xls',3, 1))

    # 输出第一行数据
    # print(excel.readTitleRow('interface.xls'))
    # 输出除了第一行以外的没行数据
    # print(excel.readDataExceptTitleRow('interface.xls'))

    # 输出第一列
    # print(excel.readTitleCol('interface.xls'))
    # 输出除第一列以外的列数据
    # print(excel.readDataExceptTitleCol('interface.xls'))

    # 读取除了第一行和第一列以外的数据
    # print(excel.readDataExceptTitleRowAndFirstCol('interface.xls'))

    # 获取该文件中的某个excel表，哪个子sheet拥有最多的行
    # print(excel.get_all_sheets_longest_rows_number('interface.xls'))

    # 获取文件夹下所有excel表下子sheet最多的行数
    # print(excel.get_tables_list())


    # print(excel.get_all_tables_longest_rows_number())

    # read_xlsx_by_openpyxl(excel_path)